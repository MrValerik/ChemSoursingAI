"""Импорт списка позиций закупки из XLSX/CSV.

Закупщик работает списком на 5–50 веществ. Ручное создание запроса на
каждую позицию сводит на нет основную экономию времени, ради которой
программу и делают.

Разбор детерминированный и целиком офлайновый: файл не уходит в LLM и не
сохраняется. Список сырья — коммерческая тайна закупщика, а срок хранения
и права доступа к загруженному файлу продуктового решения ещё не имеют;
пока его нет, самый безопасный файл — тот, которого нет на диске.

Модуль ничего не создаёт. Он отвечает на один вопрос: что именно система
прочитала в файле и с какими оговорками. Создание запросов и постановка
поиска — отдельная задача (MEET2-02).

Ошибки строк не складываются в общий отказ. Файл на 50 позиций, где
неверен один номер, обязан показать 49 готовых строк и одну проблемную с
номером строки, полем и причиной: иначе закупщик правит файл вслепую.
"""

from __future__ import annotations

import csv
import io
import re
import zipfile
from dataclasses import dataclass, field
from decimal import Decimal

from app.services.cas import is_valid_cas, normalize_cas, suggest_check_digit
from app.services.incoterms import (
    SUPPORTED_INCOTERMS,
    UnsupportedIncotermError,
    normalize_incoterm,
)
from app.services.page_facts import decimal_number, unit_key
from app.services.search_countries import SEARCH_COUNTRIES, normalize_search_country

# Ограничения входа. Список закупки — это десятки строк, а не выгрузка ERP:
# большой файл здесь означает ошибку выбора файла, а не большую закупку.
MAX_FILE_BYTES = 2 * 1024 * 1024
MAX_ROWS = 200
MAX_COLUMNS = 60

# Единицы объёма, которые понимает карточка запроса. Всё остальное либо
# приводится к ним с оговоркой, либо отклоняется — молча менять единицу
# нельзя, от неё зависит порядок величины закупки.
RFQ_UNITS = ("g", "kg", "t", "L", "mL")
_UNIT_KEY_TO_RFQ = {"g": "g", "kg": "kg", "mt": "t", "ml": "mL", "l": "L"}
# Единицы, которых в карточке нет. Пересчёт возможен и делается, но всегда
# с оговоркой: закупщик должен увидеть, что число в его файле и число в
# запросе — разные.
_UNIT_CONVERSIONS = {
    "mg": ("g", Decimal("0.001")),
    "lb": ("kg", Decimal("0.45359237")),
    "oz": ("g", Decimal("28.349523125")),
    "m3": ("L", Decimal("1000")),
}

CURRENCIES = ("USD", "EUR", "CNY", "RUB")

# Значение, начинающееся с этих символов, Excel исполняет как формулу.
# В импорте оно остаётся текстом и никогда не вычисляется.
_FORMULA_PREFIXES = ("=", "+", "-", "@")

_LIST_SPLIT_RE = re.compile(r"[;,\n]+")
_PERCENT_RE = re.compile(r"(\d+(?:[.,]\d+)?)\s*%")


# Заголовки колонок. Файл приходит от закупщика, а не из нашей выгрузки,
# поэтому один и тот же столбец называют по-разному и на двух языках.
_COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "name": (
        "название", "наименование", "вещество", "продукт", "товар", "позиция",
        "name", "substance", "product", "material", "item",
    ),
    "cas": ("cas", "cas-номер", "cas номер", "касс", "кас", "cas no", "cas number"),
    "synonyms": (
        "синонимы", "другие названия", "равнозначные названия",
        "synonyms", "aliases", "other names",
    ),
    "specification": (
        "спецификация", "требования", "описание", "характеристики",
        "specification", "spec", "requirements", "description",
    ),
    "purity": ("чистота", "purity", "assay"),
    "grade": ("грейд", "грэйд", "марка", "стандарт", "grade", "standard"),
    "volume": ("объём", "объем", "количество", "кол-во", "volume", "quantity", "qty", "amount"),
    "unit": ("единица", "ед", "ед.", "ед. изм.", "единица измерения", "unit", "uom", "measure"),
    "target_price": (
        "целевая цена", "ориентир цены", "цена", "target price", "price", "target",
    ),
    "currency": ("валюта", "currency", "curr"),
    "incoterms": (
        "incoterms", "инкотермс", "базис", "базис поставки", "условия поставки",
        "delivery terms", "delivery basis",
    ),
    "countries": ("страны", "страна", "страны поиска", "countries", "country", "origin"),
    "comment": (
        "комментарий", "примечание", "заметка", "comment", "note", "notes", "remark",
    ),
}

_ALIAS_TO_FIELD = {
    alias: field_name
    for field_name, aliases in _COLUMN_ALIASES.items()
    for alias in aliases
}


class RfqImportError(ValueError):
    """Файл не удалось прочитать целиком: разбирать нечего."""


@dataclass
class ImportIssue:
    """Замечание к строке или к файлу.

    `row` — номер строки в исходном файле, как его видит закупщик в Excel:
    заголовок это строка 1. Без номера строки замечание бесполезно —
    править файл придётся перебором.
    """

    message: str
    row: int | None = None
    column: str | None = None
    field: str | None = None

    def to_dict(self) -> dict:
        return {
            "message": self.message,
            "row": self.row,
            "column": self.column,
            "field": self.field,
        }


@dataclass
class ImportRow:
    row: int
    values: dict
    raw: dict
    errors: list[ImportIssue] = field(default_factory=list)
    warnings: list[ImportIssue] = field(default_factory=list)

    @property
    def importable(self) -> bool:
        return not self.errors

    def to_dict(self) -> dict:
        return {
            "row": self.row,
            "values": self.values,
            "raw": self.raw,
            "errors": [item.to_dict() for item in self.errors],
            "warnings": [item.to_dict() for item in self.warnings],
            "importable": self.importable,
        }


@dataclass
class ImportPreview:
    rows: list[ImportRow] = field(default_factory=list)
    file_warnings: list[ImportIssue] = field(default_factory=list)
    recognised_columns: dict[str, str] = field(default_factory=dict)
    ignored_columns: list[str] = field(default_factory=list)

    def to_dict(self) -> dict:
        return {
            "rows": [row.to_dict() for row in self.rows],
            "file_warnings": [item.to_dict() for item in self.file_warnings],
            "recognised_columns": self.recognised_columns,
            "ignored_columns": self.ignored_columns,
            "total_rows": len(self.rows),
            "importable_rows": sum(1 for row in self.rows if row.importable),
        }


def _clean(value: object) -> str:
    """Значение ячейки как строка, без вычисления формул."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "да" if value else ""
    if isinstance(value, float) and value.is_integer():
        # openpyxl отдаёт целое число как 500.0; в запрос это уходит текстом,
        # и «500.0 kg» в письме поставщику выглядит небрежностью.
        return str(int(value))
    return str(value).strip()


def _is_formula(value: str) -> bool:
    return value.startswith(_FORMULA_PREFIXES) and len(value) > 1


def _split_list(value: str) -> list[str]:
    return [item.strip() for item in _LIST_SPLIT_RE.split(value) if item.strip()]


def _read_csv(payload: bytes) -> list[list[str]]:
    """CSV с неизвестной кодировкой и неизвестным разделителем.

    Кодировку файл не сообщает, а закупщик выгружает его из чего угодно:
    Excel на русской Windows пишет cp1251, выгрузка из веба — UTF-8, часть
    редакторов добавляет BOM. Порядок попыток от строгого к терпимому;
    latin-1 в конце не падает никогда и оставляет данные читаемыми хотя бы
    частично — это лучше, чем отказ от всего файла.
    """
    text: str | None = None
    for encoding in ("utf-8-sig", "utf-8", "cp1251", "latin-1"):
        try:
            text = payload.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:  # pragma: no cover - latin-1 декодирует любой байт
        raise RfqImportError("Не удалось определить кодировку файла.")

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        # Один столбец или необычный формат: запятая — разумное умолчание,
        # и разбор всё равно даст одну колонку, которую видно в предпросмотре.
        delimiter = ";" if sample.count(";") > sample.count(",") else ","
    return [row for row in csv.reader(io.StringIO(text), delimiter=delimiter)]


def _read_xlsx(payload: bytes) -> tuple[list[list[str]], list[ImportIssue]]:
    """XLSX через openpyxl.

    `data_only=True` отдаёт посчитанное значение формулы, а не саму формулу.
    Если файл ни разу не открывали в Excel, посчитанного значения в нём нет
    и ячейка приходит пустой — это не ошибка разбора, но закупщик должен об
    этом узнать, иначе пустая колонка выглядит потерей данных.
    """
    from openpyxl import load_workbook

    warnings: list[ImportIssue] = []
    try:
        book = load_workbook(
            io.BytesIO(payload), data_only=True, read_only=True, keep_links=False
        )
    except zipfile.BadZipFile as exc:
        raise RfqImportError(
            "Файл не читается как XLSX. Старый формат .xls не поддерживается — "
            "пересохраните файл как .xlsx или .csv."
        ) from exc
    except Exception as exc:  # noqa: BLE001 - библиотека бросает разное
        raise RfqImportError(f"Файл не удалось прочитать как XLSX: {exc}") from exc

    try:
        sheet = book.worksheets[0]
        if len(book.worksheets) > 1:
            warnings.append(
                ImportIssue(
                    message=(
                        f"В книге {len(book.worksheets)} листов. Прочитан только "
                        f"первый — «{sheet.title}»."
                    )
                )
            )
        rows = [
            [_clean(cell) for cell in row]
            for row in sheet.iter_rows(max_col=MAX_COLUMNS, values_only=True)
        ]
    finally:
        book.close()

    stale = _formulas_without_values(payload, rows)
    if stale:
        # Пустая ячейка вместо формулы выглядит как потеря данных, и молчать
        # об этом нельзя: закупщик решит, что колонку не перенесли.
        shown = ", ".join(stale[:10])
        tail = f" и ещё {len(stale) - 10}" if len(stale) > 10 else ""
        warnings.append(
            ImportIssue(
                message=(
                    f"Ячейки с формулами прочитаны как пустые ({shown}{tail}). "
                    "Excel хранит посчитанное значение только после сохранения "
                    "файла — откройте его в Excel, сохраните и загрузите снова, "
                    "либо впишите значения числами."
                )
            )
        )
    return rows, warnings


def _formulas_without_values(payload: bytes, values: list[list[str]]) -> list[str]:
    """Адреса ячеек, где формула есть, а посчитанного значения нет.

    `data_only=True` отдаёт то, что Excel сохранил в файл при последнем
    закрытии. Книга, собранная программой и ни разу не открытая в Excel,
    посчитанных значений не содержит вовсе, и формульная ячейка приходит
    пустой. Отличить её от честно пустой можно только вторым проходом — по
    самим формулам.
    """
    from openpyxl import load_workbook

    try:
        book = load_workbook(
            io.BytesIO(payload), data_only=False, read_only=True, keep_links=False
        )
    except Exception:  # noqa: BLE001 - подсказка не стоит падения разбора
        return []
    stale: list[str] = []
    try:
        sheet = book.worksheets[0]
        for row_index, row in enumerate(
            sheet.iter_rows(max_col=MAX_COLUMNS), start=1
        ):
            for cell in row:
                if not isinstance(cell.value, str) or not cell.value.startswith("="):
                    continue
                read = ""
                if row_index <= len(values) and cell.column - 1 < len(
                    values[row_index - 1]
                ):
                    read = values[row_index - 1][cell.column - 1]
                if not read:
                    stale.append(cell.coordinate)
    except Exception:  # noqa: BLE001
        return stale
    finally:
        book.close()
    return stale


def _match_columns(
    header: list[str],
) -> tuple[dict[int, str], dict[str, str], list[str]]:
    """Сопоставление заголовков файла с полями запроса."""
    by_index: dict[int, str] = {}
    recognised: dict[str, str] = {}
    ignored: list[str] = []
    for index, title in enumerate(header):
        name = title.strip()
        if not name:
            continue
        field_name = _ALIAS_TO_FIELD.get(name.casefold())
        if field_name is None or field_name in recognised.values():
            # Повтор уже занятого поля не перетирает первый столбец: какой
            # из двух «Цена» правильный, знает закупщик, а не программа.
            ignored.append(name)
            continue
        by_index[index] = field_name
        recognised[name] = field_name
    return by_index, recognised, ignored


def _row_values(row: list[str], by_index: dict[int, str]) -> dict[str, str]:
    values: dict[str, str] = {}
    for index, field_name in by_index.items():
        values[field_name] = row[index].strip() if index < len(row) else ""
    return values


def _parse_cas(raw: str, row: ImportRow) -> None:
    if not raw:
        return
    if _is_formula(raw):
        row.errors.append(
            ImportIssue(
                message="Значение похоже на формулу и как номер не читается.",
                row=row.row, column="CAS", field="cas",
            )
        )
        return
    candidate = normalize_cas(raw)
    if is_valid_cas(candidate):
        row.values["cas"] = candidate
        return
    hint = suggest_check_digit(candidate)
    row.errors.append(
        ImportIssue(
            message=(
                f"CAS «{raw}» не прошёл проверку контрольной суммы. "
                + (f"Похоже, имелся в виду {hint}." if hint else "Проверьте номер.")
            ),
            row=row.row, column="CAS", field="cas",
        )
    )


def _parse_volume(raw_volume: str, raw_unit: str, row: ImportRow) -> None:
    """Объём и единица. Единицы может не быть в отдельной колонке.

    «500 кг» одной ячейкой встречается чаще, чем два аккуратных столбца,
    поэтому единица ищется и внутри значения тоже.
    """
    if not raw_volume:
        return
    if _is_formula(raw_volume):
        row.errors.append(
            ImportIssue(
                message="Значение похоже на формулу и как объём не читается.",
                row=row.row, column="Объём", field="volume",
            )
        )
        return

    text = raw_volume.strip()
    unit_text = raw_unit.strip()
    if not unit_text:
        match = re.match(r"^\s*([\d\s., ]+)\s*([^\d\s].*)$", text)
        if match:
            text, unit_text = match.group(1), match.group(2)

    amount = decimal_number(text)
    if amount is None:
        row.errors.append(
            ImportIssue(
                message=f"Объём «{raw_volume}» не читается как положительное число.",
                row=row.row, column="Объём", field="volume",
            )
        )
        return

    if not unit_text:
        row.errors.append(
            ImportIssue(
                message=(
                    f"У объёма «{raw_volume}» не указана единица измерения. "
                    "Без неё нельзя понять, идёт речь о граммах или тоннах."
                ),
                row=row.row, column="Единица", field="unit",
            )
        )
        return

    key = unit_key(unit_text)
    if key is None:
        row.errors.append(
            ImportIssue(
                message=(
                    f"Единица «{unit_text}» не распознана. "
                    f"Поддерживаются: {', '.join(RFQ_UNITS)} и их обычные написания."
                ),
                row=row.row, column="Единица", field="unit",
            )
        )
        return

    if key in _UNIT_KEY_TO_RFQ:
        unit = _UNIT_KEY_TO_RFQ[key]
    else:
        unit, factor = _UNIT_CONVERSIONS[key]
        converted = amount * factor
        row.warnings.append(
            ImportIssue(
                message=(
                    f"Единица «{unit_text}» в карточке запроса не поддерживается. "
                    f"Пересчитано: {_number_text(amount)} {unit_text} = "
                    f"{_number_text(converted)} {unit}."
                ),
                row=row.row, column="Единица", field="unit",
            )
        )
        amount = converted

    row.values["volume"] = f"{_number_text(amount)} {unit}"


def _number_text(value: Decimal) -> str:
    normalized = value.normalize()
    text = format(normalized, "f")
    return text.rstrip("0").rstrip(".") if "." in text else text


def _parse_purity(raw_purity: str, raw_grade: str, row: ImportRow) -> None:
    """Чистота и грейд лежат в карточке одной строкой: «min 99%, USP»."""
    parts: list[str] = []
    purity = raw_purity.strip()
    if purity and not _is_formula(purity):
        match = _PERCENT_RE.search(purity)
        if match:
            parts.append(f"min {match.group(1).replace(',', '.')}%")
        else:
            number = decimal_number(purity)
            if number is not None and number <= 100:
                parts.append(f"min {_number_text(number)}%")
            else:
                row.warnings.append(
                    ImportIssue(
                        message=(
                            f"Чистота «{raw_purity}» не читается как процент — "
                            "значение перенесено в требования как есть."
                        ),
                        row=row.row, column="Чистота", field="purity",
                    )
                )
                parts.append(purity)
    grade = raw_grade.strip()
    if grade and not _is_formula(grade):
        parts.append(grade)
    if parts:
        row.values["purity"] = ", ".join(parts)


def _parse_price(raw_price: str, raw_currency: str, row: ImportRow) -> None:
    price = raw_price.strip()
    if price and not _is_formula(price):
        amount = decimal_number(price)
        if amount is None:
            row.warnings.append(
                ImportIssue(
                    message=(
                        f"Ориентир цены «{raw_price}» не читается как "
                        "положительное число и не перенесён."
                    ),
                    row=row.row, column="Целевая цена", field="target_price",
                )
            )
        else:
            row.values["target_price"] = float(amount)

    currency = raw_currency.strip().upper()
    if currency:
        if currency in CURRENCIES:
            row.values["currency"] = currency
        else:
            row.warnings.append(
                ImportIssue(
                    message=(
                        f"Валюта «{raw_currency}» не поддерживается. "
                        f"Оставлен USD. Доступны: {', '.join(CURRENCIES)}."
                    ),
                    row=row.row, column="Валюта", field="currency",
                )
            )


def _parse_incoterms(raw: str, row: ImportRow) -> None:
    if not raw.strip():
        return
    codes: list[str] = []
    for item in _split_list(raw):
        try:
            code = normalize_incoterm(item)
        except UnsupportedIncotermError:
            row.warnings.append(
                ImportIssue(
                    message=(
                        f"Базис поставки «{item}» не поддерживается и пропущен. "
                        f"Доступны: {', '.join(SUPPORTED_INCOTERMS)}."
                    ),
                    row=row.row, column="Incoterms", field="incoterms",
                )
            )
            continue
        if code not in codes:
            codes.append(code)
    if codes:
        row.values["incoterms"] = codes


def _parse_countries(raw: str, row: ImportRow) -> None:
    if not raw.strip():
        return
    countries: list[str] = []
    for item in _split_list(raw):
        try:
            country = normalize_search_country(item)
        except ValueError:
            row.warnings.append(
                ImportIssue(
                    message=(
                        f"Страна поиска «{item}» не поддерживается и пропущена. "
                        f"Доступны: {', '.join(SEARCH_COUNTRIES)}."
                    ),
                    row=row.row, column="Страны", field="search_countries",
                )
            )
            continue
        if country not in countries:
            countries.append(country)
    if countries:
        row.values["search_countries"] = countries[:3]


def _parse_row(row_number: int, raw: dict[str, str]) -> ImportRow:
    row = ImportRow(row=row_number, values={}, raw=dict(raw))

    name = raw.get("name", "").strip()
    if _is_formula(name):
        row.errors.append(
            ImportIssue(
                message="Название похоже на формулу и как название не читается.",
                row=row_number, column="Название", field="name",
            )
        )
    elif not name:
        row.errors.append(
            ImportIssue(
                message="Не указано название — по нему идёт поиск, без него строка бесполезна.",
                row=row_number, column="Название", field="name",
            )
        )
    else:
        row.values["name"] = name

    _parse_cas(raw.get("cas", "").strip(), row)
    _parse_volume(raw.get("volume", ""), raw.get("unit", ""), row)
    _parse_purity(raw.get("purity", ""), raw.get("grade", ""), row)
    _parse_price(raw.get("target_price", ""), raw.get("currency", ""), row)
    _parse_incoterms(raw.get("incoterms", ""), row)
    _parse_countries(raw.get("countries", ""), row)

    synonyms = [
        item for item in _split_list(raw.get("synonyms", "")) if not _is_formula(item)
    ]
    if synonyms:
        row.values["confirmed_synonyms"] = synonyms[:50]

    specification = raw.get("specification", "").strip()
    if specification and not _is_formula(specification):
        row.values["specification"] = specification[:4000]

    comment = raw.get("comment", "").strip()
    if comment and not _is_formula(comment):
        row.values["specialist_comment"] = comment[:4000]

    # Способ идентификации выводится из заполненного, как и в форме.
    row.values["identification_method"] = "cas" if row.values.get("cas") else "spec"
    return row


def parse_import_row(row_number: int, raw: dict[str, str]) -> ImportRow:
    """Пересобирает одну строку из исправленных закупщиком значений.

    Правка в предпросмотре обязана проходить ту же проверку, что и разбор
    файла. Иначе исправленное значение живёт по другим правилам, чем
    прочитанное, и в запрос уходит то, что парсер бы не пропустил.
    """
    cleaned = {
        key: str(value or "").strip()
        for key, value in raw.items()
        if key in _COLUMN_ALIASES
    }
    return _parse_row(row_number, cleaned)


def parse_import_file(filename: str, payload: bytes) -> ImportPreview:
    """Разбирает файл и возвращает предпросмотр. Ничего не сохраняет."""
    if not payload:
        raise RfqImportError("Файл пустой.")
    if len(payload) > MAX_FILE_BYTES:
        raise RfqImportError(
            f"Файл больше {MAX_FILE_BYTES // (1024 * 1024)} МБ. "
            "Список закупки — это десятки строк, а не выгрузка из учётной системы."
        )

    lowered = (filename or "").casefold()
    preview = ImportPreview()
    if lowered.endswith(".xlsx") or lowered.endswith(".xlsm"):
        table, warnings = _read_xlsx(payload)
        preview.file_warnings.extend(warnings)
    elif lowered.endswith(".csv") or lowered.endswith(".txt"):
        table = _read_csv(payload)
    else:
        raise RfqImportError(
            "Поддерживаются только .xlsx и .csv. Формат определяется по "
            "расширению имени файла."
        )

    # Пустые строки перед заголовком — обычное дело: в файле бывает шапка
    # с названием отдела или пустая строка сверху.
    header_index = next(
        (i for i, line in enumerate(table) if any(cell.strip() for cell in line)),
        None,
    )
    if header_index is None:
        raise RfqImportError("В файле нет ни одной непустой строки.")

    header = table[header_index]
    by_index, recognised, ignored = _match_columns(header)
    preview.recognised_columns = recognised
    preview.ignored_columns = ignored

    if "name" not in by_index.values():
        raise RfqImportError(
            "В файле не найдена колонка с названием вещества. Ожидается "
            "заголовок «Название» или «Name». Найдены: "
            + (", ".join(item for item in header if item.strip()) or "пусто")
        )
    if ignored:
        # Молча выбрасывать колонку нельзя: закупщик считает, что данные
        # перенесены, а их нет.
        preview.file_warnings.append(
            ImportIssue(
                message=(
                    "Колонки не распознаны и пропущены: "
                    + ", ".join(f"«{item}»" for item in ignored)
                    + ". Данные из них в запросы не попадут."
                )
            )
        )

    body = table[header_index + 1 :]
    if len(body) > MAX_ROWS:
        preview.file_warnings.append(
            ImportIssue(
                message=(
                    f"В файле {len(body)} строк, прочитаны первые {MAX_ROWS}. "
                    "Разделите список на несколько файлов."
                )
            )
        )
        body = body[:MAX_ROWS]

    seen: dict[tuple[str, str], int] = {}
    for offset, line in enumerate(body):
        if not any(cell.strip() for cell in line):
            # Пустая строка внутри файла — разделитель, а не ошибка.
            continue
        # Номер строки такой же, как в Excel: заголовок был строкой
        # header_index + 1, значит тело начинается со следующей.
        row_number = header_index + offset + 2
        raw = _row_values(line, by_index)
        row = _parse_row(row_number, raw)

        key = (
            row.values.get("name", "").casefold(),
            row.values.get("cas", ""),
        )
        if key != ("", "") and key in seen:
            row.warnings.append(
                ImportIssue(
                    message=(
                        f"Такая же позиция уже есть в строке {seen[key]}. "
                        "Запросы создадутся оба, если строку не исключить."
                    ),
                    row=row_number, column="Название", field="name",
                )
            )
        elif key != ("", ""):
            seen[key] = row_number

        preview.rows.append(row)

    if not preview.rows:
        raise RfqImportError("В файле нет строк с данными — только заголовок.")
    return preview


# --- Образец файла ---
#
# Закупщик, открывший экран впервые, не знает, в каком виде нужен файл.
# Перечисление колонок в подсказке этого не решает: «Объём» и «Единица»
# как заголовки понятны, а вот что писать в «Чистота» и как перечислить
# два базиса поставки — уже нет. Заполненный образец отвечает на это
# показом, а не описанием.
#
# Заголовки образца обязаны читаться тем же разбором, что и чужой файл:
# скачанный образец, залитый обратно без правок, должен дать готовые
# строки. За этим следит тест — он гоняет образец через parse_import_file.

TEMPLATE_COLUMNS: tuple[tuple[str, str, str], ...] = (
    (
        "name",
        "Название",
        "Обязательна. Название, марка или торговое наименование — так, "
        "как его пишут поставщики. По нему идёт поиск.",
    ),
    (
        "cas",
        "CAS",
        "Если номер известен. Проверяется контрольная сумма: неверный "
        "номер строку не пропустит. Для смесей и марок номера нет — "
        "оставьте пусто и заполните «Спецификация».",
    ),
    (
        "synonyms",
        "Синонимы",
        "Равнозначные названия через запятую или точку с запятой. "
        "Например: Ascorbic acid; Витамин C.",
    ),
    (
        "specification",
        "Спецификация",
        "Требования, по которым вещество опознаётся без номера: "
        "показатели, диапазоны, назначение.",
    ),
    (
        "purity",
        "Чистота",
        "Процент: 99, 99% или 98,5%. Непроцентное значение перенесётся "
        "в требования как есть.",
    ),
    ("grade", "Грейд", "Марка или стандарт: USP, BP, food grade, tech."),
    (
        "volume",
        "Объём",
        "Число. Единица — в соседней колонке; «500 кг» одной ячейкой "
        "тоже читается.",
    ),
    (
        "unit",
        "Единица",
        f"Одна из: {', '.join(RFQ_UNITS)} или их русские написания "
        "(г, кг, т, л, мл). mg, lb, oz, m3 пересчитываются с оговоркой.",
    ),
    (
        "target_price",
        "Целевая цена",
        "Ориентир за единицу. Поставщику не показывается — нужен для "
        "сравнения предложений.",
    ),
    ("currency", "Валюта", f"Одна из: {', '.join(CURRENCIES)}. Пусто — USD."),
    (
        "incoterms",
        "Incoterms",
        f"Базисы поставки через запятую: {', '.join(SUPPORTED_INCOTERMS)}. "
        "Пусто — возьмётся общий набор, отмеченный на экране под таблицей.",
    ),
    (
        "countries",
        "Страны",
        f"Страны поиска через запятую: {', '.join(SEARCH_COUNTRIES)}. "
        "Пусто — возьмутся общие, отмеченные на экране под таблицей.",
    ),
    (
        "comment",
        "Комментарий",
        "Внутренняя заметка закупщика. В письмо поставщику не уходит.",
    ),
)

# Строки образца. Три случая, которые закрывают почти весь список:
# позиция с номером, позиция без номера (её опознают по спецификации) и
# позиция, заполненная по минимуму. Данные демонстрационные.
TEMPLATE_ROWS: tuple[dict[str, str], ...] = (
    {
        "name": "Аскорбиновая кислота",
        "cas": "50-81-7",
        "synonyms": "Ascorbic acid; Витамин C",
        "purity": "99%",
        "grade": "USP",
        "volume": "500",
        "unit": "кг",
        "target_price": "6.5",
        "currency": "USD",
        "incoterms": "CIP, FCA",
        "countries": "Китай, Индия",
        "comment": "Нужен паспорт качества на партию.",
    },
    {
        "name": "Полиэфирполиол для жёсткого пенополиуретана",
        "synonyms": "Polyether polyol",
        "specification": (
            "Гидроксильное число 440-460 мг KOH/г, вязкость "
            "8000-12000 мПа·с при 25 °C, функциональность 4,3"
        ),
        "volume": "20",
        "unit": "т",
        "incoterms": "FCA",
        "countries": "Китай",
        "comment": "Номера CAS нет — искать по спецификации.",
    },
    {
        "name": "Глицин",
        "cas": "56-40-6",
        "volume": "2",
        "unit": "т",
    },
)

# Выбор единицы в Excel. Русские написания идут первыми: закупщик пишет
# «кг», и список, где их нет, выглядел бы запретом на привычную запись —
# хотя разбор её понимает.
TEMPLATE_UNIT_CHOICES = ("кг", "т", "г", "л", "мл", *RFQ_UNITS)

TEMPLATE_FORMATS = ("xlsx", "csv")

TEMPLATE_MEDIA_TYPES = {
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "csv": "text/csv; charset=utf-8",
}


def _template_table() -> list[list[str]]:
    header = [title for _, title, _ in TEMPLATE_COLUMNS]
    body = [
        [row.get(field_name, "") for field_name, _, _ in TEMPLATE_COLUMNS]
        for row in TEMPLATE_ROWS
    ]
    return [header, *body]


def build_template_csv() -> bytes:
    """Образец в CSV.

    Разделитель — точка с запятой, кодировка — UTF-8 с BOM: Excel на
    русской Windows открывает такой файл разложенным по столбцам, а файл
    с запятой — одной склеенной колонкой. Наш разбор понимает оба.
    """
    buffer = io.StringIO()
    csv.writer(buffer, delimiter=";", lineterminator="\r\n").writerows(
        _template_table()
    )
    return buffer.getvalue().encode("utf-8-sig")


def build_template_xlsx() -> bytes:
    """Образец в XLSX: заголовки с пояснениями и три заполненные строки.

    Пояснение к каждой колонке живёт примечанием на её заголовке, а не
    отдельным листом: лист с инструкцией разбор бы не прочитал и честно
    предупредил бы, что взял только первый, — а закупщик увидел бы это
    как ошибку.
    """
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    table = _template_table()
    book = Workbook()
    sheet = book.active
    sheet.title = "Позиции"
    for line in table:
        sheet.append(line)

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="EEF2F7")
    # Ширины по смыслу колонки: спецификация и комментарий длинные, номер
    # и единица короткие. Без этого образец открывается частоколом «###».
    widths = {
        "name": 34,
        "cas": 13,
        "synonyms": 26,
        "specification": 46,
        "purity": 11,
        "grade": 12,
        "volume": 9,
        "unit": 10,
        "target_price": 13,
        "currency": 9,
        "incoterms": 16,
        "countries": 16,
        "comment": 34,
    }
    for index, (field_name, title, hint) in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
        cell.comment = Comment(
            f"{title}\n{hint}", "ChemSource AI", height=150, width=340
        )
        sheet.column_dimensions[cell.column_letter].width = widths[field_name]

    sheet.freeze_panes = "A2"

    # Списки выбора там, где значение одно и набор закрытый. Для базисов
    # и стран списка нет намеренно: там перечисление через запятую, и
    # проверка Excel отвергала бы верное «CIP, FCA».
    last_row = len(table) + 40
    choices = (("unit", TEMPLATE_UNIT_CHOICES), ("currency", CURRENCIES))
    for field_name, allowed in choices:
        index = next(
            position
            for position, (name, _, _) in enumerate(TEMPLATE_COLUMNS, start=1)
            if name == field_name
        )
        letter = sheet.cell(row=1, column=index).column_letter
        rule = DataValidation(
            type="list",
            formula1='"' + ",".join(allowed) + '"',
            allow_blank=True,
        )
        sheet.add_data_validation(rule)
        rule.add(f"{letter}2:{letter}{last_row}")

    stream = io.BytesIO()
    book.save(stream)
    return stream.getvalue()
